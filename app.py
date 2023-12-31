"""
    1. PRINT FUNCTIONS
   output => print
"""
# print("Hezron K.")
# string & Expression
# print('*' * 10)

""" 
    2. VARIABLE => is temporary store  the data in computer memory, for example:
    # value will be converted to binary (0's & 1's)
    # integer and floats
    # Python is case-sensitive 
"""
price = 10
rating = 4.9
# string
name = 'Hezron'
# Boolean
is_published = False
# print(price)

# Receive User Input
# name = input('What is Your name? ')
# concantination
# print = ('Hi' + name)


"""
    3. TYPE CONVERSION
"""
#  user input is a string therefore we intend to convert it to integer

# Functions :=> bool() => convert a string into a boolean(True or False) float() => convert a string into a float, int() => convert a string into integer, type() => to check the type for a particular datatype

# birth_yr = input('Birth Year: ')
# age = 2023 - int(birth_yr)

# print(age)


"""
    4. strings => single quotes and double quotes.

"""

# course = "Python for beginners"
# print(course[1]) # RESULTS => P
# range
# print(course[0:3]) # RESULTS => Pyt
# print(course[:5]) # RESULTS => Pytho

name = "Hezron"
# print(name[1:-1]) # RESULTS => ezro

"""
    FORMATTED STRINGS
"""

first = 'Hezron'
last = 'smith'

message = first + '[' + last + '] is a code'
# print(message)

# Below is the formwtted string
msg = f'{first} [{last}] is a code'
# print(msg)

"""
     String Methods e.g len
"""
course = "Python for beginners"

# print(len(course))
# print(course.upper())
# print(course.lower())
# print(course.find('P'))


"""
    Arithmetic e.g / % - +  ** 
"""

# assignment operator
x = 10
x = x + 3 # OR x += 3
# print(x)

# operator precendence 
#  ======= BODMAS ========

"""
    1. Exponentiation 2 ** 3
    2. Multiplication or Division
    3. Addition or Substraction
    4. Add Parenthesis will take place first
"""

y = 10 + 3 * 2 ** 2
# print(y)

"""
    Math functions e.g round(), abs() returns a positive value, modules => import math
"""

import math
z = 2.9
# print(round(z))
# print(math.ceil(2.9))
# print(math.floor(2.9))

"""
    # IF STATEMENT e.g  if it's hot 
                            It's a hot day
                            Drink plenty of water
                    Otherwise if it's cold
                        it's a cold day 
                        Wear warm clothes
                Otherwise
                    It's a lovely day

"""

is_hot = False
is_cold = True

# if is_hot:
#     print("It's a hot day")
#     print("Drink plenty of water")
# elif is_cold:
#     print("It's a cold day")
#     print("Wear warm cloth")
     
# else:
#     print("Enjoy your day")

"""
    Logical operators  e.g  OR, AND , NOT
"""
# has_high_income = True
# has_good_credit = True

# if has_good_credit and has_good_credit:
#     print("El,igible for a loan")

"""
    Comparison operators >, >=, <=, ==,!=
"""

# weight = int(input("Weight: "))
# unit = input('(L)bs or (K)g: ')

# if unit.upper() == "L":
#     converted = weight * 0.45
#     print(f"You'r {converted} kilos")
# else:
#     converted = weight / 0.45
#     print(f"You're {converted} pounds")

"""
    While Loops
# """
# # initialize or asign i 1
# i = 1
# while i<= 5:
#     print("*" * i)
#     # Increment i
#     i += 1
# print("Done")

"""
    For loops => iteralate over character in a string
    1. list, characters
"""

# for item in "Python":
#     print(item)
# # output: loops through the word e.g P, y, t, h, o, n

# for item in [1, 2, 3, 4, 5]:
#     print(item)

# range => create a object., which can iteralate over
# for item in range(10):
#     print(item)

# Output: 0, 1, 2, 3, 4, 5, 6, 7, 8, 9

# for item in range(5, 10):
#     print(item)

# # output: start from 5, 6, 7, 8, 9

# for item in range(5, 10, 2):
#     print(item)

# # output: 5, then goes 2 step forward, 7, then goes 2 step forward print 9

# #  output: 5, 7, 9

prices = [10, 20, 30]
total = 0

for price in prices:
    total += price

# print(f"Total: {total}")

"""
    nested loop
"""

# for x in range(4):
#     for y in range(3):
#         print(f'({x}, {y})')


"""
OUTPUT::
XXXXX
XX
XXXXX
XX
XX
"""
numbers = [2, 2, 2, 2, 15]

# for x in numbers:
#     print(f'x' * x)

for x_count in numbers:
    output = ''
    for count in range(x_count):
        output += 'X'
    # print(output)


    """
        LIST => it's array which can hold different data types such as integer, strings, float
    """
n = ["John", "Allan", "Hezron", "Baraka"]
# print(n)
# accessing individual in a list
# print(n[-1])
# access in a range
# print(n[:3])


"""
    2D LISTS => are stream powerful and have alot application in data science  and machine language.
    It's a rectangle array of elements which have rows and column 
"""
# matrix list
matrix = [
    # 1st list
    [1, 2, 3],
    # 2nd list
    [4, 5, 6],
    # 3rd list
    [7, 8, 9]
]

# access

# print(matrix[0])
# # loop through the 2D list
# for row in matrix:
#     for item in row:
#         print(item)

# modifying element in the 2D list
# matrix[1][1] = 0
# print(matrix)


"""
    LIST METHODS
"""

numbers = [5, 2, 1, 7, 4]

# numbers.append(13)
# numbers.insert(0, 10)
# numbers.remove(5)
# numbers.clear()
# numbers.pop()
# numbers.sort()
# return no. of elements are the same
# numbers.count()
# numbers.reverse()
# copy of numbers list
# numbers2 = numbers.copy()
# print(numbers.index(1))
# print(numbers)

# print(50 in numbers)

"""
    TUPLES => are immutable
"""

numbers = (1, 2, 3)

# methods
# numbers.count()
# numbers.index()

"""
    Unpacking
"""

coordinates = (1, 2, 3) 

# print(coordinates[0] * coordinates[1] * coordinates[2])

# # OR
# x = coordinates[0]
# y = coordinates[1]
# z = coordinates[2]

# print(x * y * z)

# On packing, the python intepret will check 1st element and assign to x, check the next element and assign to y and last element and assign to z.

# NOTED :: UNPACKING IS NOT LIMITED TO TUPLES ONLY BUT YOU CAN USE IT TO LIST ALSO.
x, y, z = coordinates

# print(x, z, y)


"""
    Dictionary
"""

# think that customer has attributes like name, age, phone, email

# E.G

"""
    name : John
    email: hezzy@kialo.com
    phone: +25470000000

    Dictionary has keys and values, therefore, in my case the keys are name, email & phone and value are John, hezzy@kialo.com & 25470000000 respectively
"""

# Define a Dictionary

customer = {
    "name": "John",
    "age": 30,
    "is_verified": True
}

# access by keyvalue
# print(customer["name"])
# use get method => return none value becoz  the key doesn't exist
# print(customer.get("bith"))

"""
    function => build large program, we need to break down
"""
"""
    parameters => are place holder
"""
# def greet_user(first_name, last_name):
#     print(f'Hi there {first_name} {last_name}')
# arqument => position arqument

# print(greet_user("hezron", "ngoma"))

# key argument
# print(greet_user(last_name="hezron", first_name="ngoma"))

# return statement
def square(num):
    return num * num


# print(square(3))

"""
    EXCEPTIONS
"""
# try: 
#     age = int(input('Age: '))
#     income = 2000
#     risk = income / age
#     print(age)
# except ZeroDivisionError:
#     print("Age Cann't be Zero")
    
# except ValueError:
#     print('Invalid Value')
    

"""
    classes => is custom datatypes and blueprint of a object.
    ================================= BASIC TYPES =============================================
    1. Numbers
    2. Strings
    3. Booleans

    ================================= Complex Types =============================================
    1. List
    2. Dictionaries
"""

# Use class keyword and name-of-you-class and shld start with Capital Letter and re
class Point:
    # construtor
    def __init__(self, x, y):
        self.x = x
        self.y = y

    def move(self):
        print("Move")

    def draw(self):
        print("draw")

# object is a instance of a class

# point1 = Point()
# point1.x = 10
# point1.y = 20
# print(point1.y)
# point1.draw()


# point2 = Point()
# print(point2)

"""]
    Module used to organize our code
"""

import converts
# print(converts.kg_to_lbs(7))


# Package

import ecommerce.shipping
from  ecommerce.shipping import calc_shipping

# ecommerce.shipping.calc_shipping()

# calc_shipping()

"""
    Use Inbuilt module
"""

import random

# for i in range(3):
#     print(random.randint(10, 20))


# list 

members = ['John', 'Mary', 'Bob', 'Hezron']
# print(random.choice(members))


"""
    files and Directories
"""

from pathlib import Path

# Path() => is a class and it will reference the current directory

path = Path("OOP")
# print(path.exists())  
print(path.mkdir())
# print(path.rmkdir())
# print(path.glob('*.py'))

# for file in path.glob('*.py'):
#     print(file)

"""
    Automation with python 
"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell
    # cell = sheet['a1']
    # OR
    cell = sheet.cell(1, 1)
    # print(cell.value)
    # # check the no. of rows
    # print(sheet.max_row)
    # # check the no. of column
    # print(sheet.max_column)

    # iterate
    for row in range(2, sheet.max_row +1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4) 
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.
                    max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)

    sheet.add_chart(chart, 'e2')

    # print(wb.save(filename))


"""
    WHAT'S MACHINE LEARNING 
    1. Self-driving Cars
    2. Robotics
    3. Language Processing
    4. Vision Processing
    5. Forecasting Stock Market Trends

    ================ Involves Number of steps =================
    STEP 1 - Import the Data => store into csv files
    STEP 2 - Clean the Data => input data is clean || depends with the data we are working with
    STEP 3 -  Split the Data into segments[Traning / Test set] e.g 80% for training and 20% for testing
    STEP 4 - Create a model - selecting a algorithm to analyse the data. the algorithms include:- Decesion trees, Neurol networks
    STEP 5 - Train the model
    STEP 6 - Make Predictions
    STEP 7 - Evaluate and improve
"""


"""
    Libraries and Tools

    ================================== LIBRARIES ==================================
    1. NUMPY => PROVIDE A MULTI-DIMENISION ARRAY
    2. PANDAS => DATA ANALYSIS THAT PROVIDE A CONCEPT CALL DATA FRAME. DATA FRAME IS A TWO DIMESION STRUCTURE SIMILARY EXCEL [ROWS, COLUMNS]
    3. MATPLOTLIB => TWO DIMESION FOR CREATING MAP AND PLOTS
    4. SCIKIT-LEARN

"""


# Import a Data Set
import pandas as pd

# df = pd.read_csv('PS4_GamesSales.csv')

# print(df)

# Useful methods and Attributes in pandas :: GET more in pandas documentation


# Create new directory
from pathlib import Path

path = Path("Intermediate")

# path.mkdir() #then execute You shall see new directory call Intermediate